from django.shortcuts import render
from django.http import  HttpResponse

# Create your views here.
# import xlrd
from . import models
from openpyxl import Workbook,load_workbook
from openpyxl.compat import range
from django.core import serializers


#######################


from django.shortcuts import render,render_to_response
from django.http import HttpResponse,HttpResponseRedirect
from django.template import RequestContext
from django import forms
from . models import Teacher, Student, ExerciseHistory

#表单
class UserForm(forms.Form):
    username = forms.CharField(label='用户名',max_length=100)
    password = forms.CharField(label='密码',widget=forms.PasswordInput())

# 学生注册
def student_regist(request):

    if request.method == 'POST':
        # 获得表单数据
        student_number = request.POST['student_number']
        password = request.POST['password']

        # 判断数据库中是否已经有此用户名
        student = Student.objects.filter(student_number__exact = student_number)
        if not student:
            # 没有此用户则添加到数据库
            Student.objects.create(student_number=student_number, password=password)
            result = models.Student.objects.all()
            json_data = serializers.serialize("json", result)
            return HttpResponse(json_data, content_type="application/json")


    result = []
    json_data = serializers.serialize("json", result)
    return HttpResponse(json_data, content_type="application/json")

# 学生登陆
def student_login(request):

    if request.method == 'POST':
        # 获得表单数据
        student_number = request.POST['student_number']
        password = request.POST['password']

        # 判断账户密码是否正确
        student = Student.objects.filter(student_number__exact = student_number, password__exact= password)
        if student:
            result = models.Student.objects.all()
            json_data = serializers.serialize("json", result)
            return HttpResponse(json_data, content_type="application/json")


    result = []  # 如果不能正确匹配此学生对应的用户名和密码，则返回空 json
    json_data = serializers.serialize("json", result)
    return HttpResponse(json_data, content_type="application/json")

# 存储上传的学生做题记录
def save_exercise_history(request):
    if request.method == 'POST':
        # 获取请求参数
        student_number = request.POST['student_number']
        question_numbers = request.POST['question_numbers']
        submit_options = request.POST['submit_options']
        score = request.POST['score']
        subject = request.POST['subject']
        # 存储
        ExerciseHistory.objects.create(student_number=student_number, subject=subject, question_numbers=question_numbers, submit_options=submit_options, score=score)
        # 返回数据
        result = models.ExerciseHistory.objects.all()
        json_data = serializers.serialize("json", result)
        return HttpResponse(json_data, content_type="application/json")

    result = []
    json_data = serializers.serialize("json", result)
    return HttpResponse(json_data, content_type="application/json")

# 返回某个学生做题记录
def get_exercise_history(request):
    if request.method == 'POST':
        # 获取请求参数
        student_number = request.POST['student_number']

        # 查找
        history = models.ExerciseHistory.objects.filter(student_number__exact=student_number).order_by('-id')
        # 如果找到记录
        if history:
            # 返回数据
            json_data = serializers.serialize("json", history)
            return HttpResponse(json_data, content_type="application/json")

    result = []
    json_data = serializers.serialize("json", result)
    return HttpResponse(json_data, content_type="application/json")

# 返回某个学生的个人信息
def get_student_profile(request):
    if request.method == 'GET':
        # 获取请求参数
        student_number = request.GET['student_number']

        # 查找
        student = models.Student.objects.filter(student_number__exact=student_number)
        # 如果找到记录
        if student:
            # 返回数据
            json_data = serializers.serialize("json", student)
            return HttpResponse(json_data, content_type="application/json")

    result = []
    json_data = serializers.serialize("json", result)
    return HttpResponse(json_data, content_type="application/json")

# 更新某个学生提交的个人信息
def update_student_profile(request):
    if request.method == 'POST':
        # 获取请求参数
        student_number = request.POST['student_number']
        name = request.POST['name']
        school = request.POST['school']
        major = request.POST['major']
        attendance_year = request.POST['attendance_year']
        password = request.POST['password']

        # 查找
        student = models.Student.objects.get(student_number__exact=student_number)
        # 如果找到记录
        if student:
            # 更新
            student.name = name
            student.school = school
            student.major = major
            student.attendance_year = attendance_year
            student.password = password
            student.save()

            # 返回更新后的数据
            student = models.Student.objects.filter(student_number__exact=student_number)
            json_data = serializers.serialize("json", student)
            return HttpResponse(json_data, content_type="application/json")

    result = []
    json_data = serializers.serialize("json", result)
    return HttpResponse(json_data, content_type="application/json")

# 教师注册
def teacher_regist(req):
    if req.method == 'POST':
        uf = UserForm(req.POST)
        if uf.is_valid():
            #获得表单数据
            username = uf.cleaned_data['username']
            password = uf.cleaned_data['password']

            # 判断数据库中是否已经有此用户名
            user = Teacher.objects.filter(username__exact = username)
            if not user:
                # 没有此用户则添加到数据库
                Teacher.objects.create(username=username, password=password)
                return render(req, 'question/teacher_login.html',{'uf':uf})

    else:
        uf = UserForm()

    # return render_to_response('question/regist.html',{'uf':uf}, context_instance=RequestContext(req))
    return render(req, 'question/teacher_regist.html',{'uf':uf})

# 教师登陆
def teacher_login(req):
    if req.method == 'POST':
        uf = UserForm(req.POST)
        if uf.is_valid():
            #获取表单用户密码
            username = uf.cleaned_data['username']
            password = uf.cleaned_data['password']
            #获取的表单数据与数据库进行比较
            user = Teacher.objects.filter(username__exact = username,password__exact = password)
            if user:
                #比较成功，跳转index
                response = HttpResponseRedirect('/question/upload_questions/')
                #将username写入浏览器cookie,失效时间为3600
                response.set_cookie('username',username,3600)
                return response
            else:
                #比较失败，还在login
                return HttpResponseRedirect('/question/teacher_login/')
    else:
        uf = UserForm()
    # return render_to_response('question/login.html',{'uf':uf},context_instance=RequestContext(req))
    return render(req, 'question/teacher_login.html',{'uf':uf})


# 教师退出
def teacher_logout(req):
    # response = HttpResponse('logout !!')
    # #清理cookie里保存username
    # response.delete_cookie('username')
    return teacher_login(req)



# 教师上传题库
def upload_questions(request):
    # return render(request, 'question/upload_questions.html')
    username = request.COOKIES.get('username', '')
    return render_to_response('question/upload_questions.html', {'username': username})

def submit_excel_action(request):

    file = request.FILES['excelFile']

    url = [] # 声明 list
    workbook = load_workbook(filename=file)
    sheetnames = workbook.get_sheet_names()
    sheet = workbook.get_sheet_by_name(sheetnames[0])

    for rowNum in range(2, sheet.max_row + 1):
        # for colNum in range(1, sheet.max_column + 1):
        title = sheet.cell(row=rowNum, column=1).value
        options = sheet.cell(row=rowNum, column=2).value
        answer = sheet.cell(row=rowNum, column=3).value
        subject_number = sheet.cell(row=rowNum, column=4).value
        subject = sheet.cell(row=rowNum, column=5).value
        chapter_number = sheet.cell(row=rowNum, column=6).value
        chapter = sheet.cell(row=rowNum, column=7).value
        difficulty_degree = sheet.cell(row=rowNum, column=8).value
        pub_time = sheet.cell(row=rowNum, column=9).value

        question = models.Question.objects.create(title=title, options=options, answer=answer, subject_number=subject_number,
                                                  subject=subject, chapter_number = chapter_number, chapter = chapter,
                                                  difficulty_degree=difficulty_degree, pub_time=pub_time)



    # all = models.Question.objects.values()
    # return HttpResponse(JSONResponse({"res": "success", "msg": all}))

    # return HttpResponse(json.dumps(data), content_type='application/json')

    json_data = serializers.serialize("json", models.Question.objects.all())

    return HttpResponse(json_data, content_type="application/json")

def get_questions_json(request):

    # 获取属性
    dd = request.GET['difficulty_degree']
    sn = request.GET['subject_number']
    cn = request.GET['chapter_number']
    isR = request.GET['isRandom']
    isN = request.GET['isNewest']
    count = request.GET['count']
    count = int(count)

    # 根据题目难度，1 为简单，2 为中等, 3 为困难, 0 为不根据难度出题
    if dd == '1':
        result = models.Question.objects.filter(subject_number=str(sn)).order_by('difficulty_degree')[:count]
    elif dd == '2':
        result = models.Question.objects.filter(subject_number=str(sn), difficulty_degree__in=[4,5,6])[:count] # 选择难度从  3-7 的题目
    elif dd == '3':
        result = models.Question.objects.filter(subject_number=str(sn)).order_by('-difficulty_degree')[:count]  # 倒序
    else:
        # 根据章节
        if cn != '0':
            result = models.Question.objects.filter(subject_number=str(sn), chapter_number=str(cn))[:count]
        # 最新题目
        elif isN == '1':
            result = models.Question.objects.filter(subject_number=str(sn)).order_by('-nid')[:count]
        # 随机出题
        elif isR == '1':
            result = models.Question.objects.filter(subject_number=str(sn)).order_by('?')[:count]
        # 缺省情况
        else:
            result = models.Question.objects.filter(subject_number=str(sn))[:10]






    json_data = serializers.serialize("json", result)

    return HttpResponse(json_data, content_type="application/json")

# 返回一些题目
def get_questions_json_by_q_nums(request):
    if request.method == 'POST':
        # 获取请求参数
        q_nums_str = request.POST['q_nums']
        q_nums = q_nums_str.split(',')
        q_nums.pop()
        # 查找
        result = models.Question.objects.filter(nid__in=q_nums)

        # 如果找到记录
        if result:
            # 返回数据
            json_data = serializers.serialize("json", result)
            return HttpResponse(json_data, content_type="application/json")

    result = []
    json_data = serializers.serialize("json", result)
    return HttpResponse(json_data, content_type="application/json")